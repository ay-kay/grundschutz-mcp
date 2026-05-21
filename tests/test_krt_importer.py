"""Tests for the KRT (Kreuzreferenztabelle) importer.

The KRT XLSX is the BSI's machine-readable mapping of requirements to
elementary threats, plus per-requirement protection-goal markers
(C/I/A). Our importer must reproduce its content verbatim into the
``requirement_threat`` and ``requirement_threat_goal`` tables.

We synthesise a tiny KRT XLSX with openpyxl so the unit tests are
deterministic; the smoketest below additionally runs against the real
krt2023.xlsx if it has been fetched via ``make fetch-krt``.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from grundschutz_mcp.ingest.krt_importer import _parse_cia, import_krt

REAL_KRT = Path(__file__).resolve().parents[1] / "krt2023.xlsx"


# ---------------------------------------------------------------------------
# CIA-string parser
# ---------------------------------------------------------------------------


def test_parse_cia_handles_combinations():
    assert _parse_cia("CIA") == ["C", "I", "A"]
    assert _parse_cia("CI") == ["C", "I"]
    assert _parse_cia("I") == ["I"]
    assert _parse_cia("AIC") == ["C", "I", "A"]  # any order tolerated


def test_parse_cia_treats_dash_and_blank_as_empty():
    assert _parse_cia("-") == []
    assert _parse_cia("") == []
    assert _parse_cia(None) == []
    assert _parse_cia("   ") == []


# ---------------------------------------------------------------------------
# Synthetic-fixture import test
# ---------------------------------------------------------------------------


def _build_synthetic_krt(path: Path) -> None:
    """Write a tiny KRT XLSX shaped like the BSI's real file.

    One sheet for the tiny_db's APP.1.1 Baustein. Links APP.1.1.A2 to
    G 0.39 (no goals), APP.1.1.A1 (deprecated) to nothing.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    # First sheet keeps its default name; rename to BSI's convention.
    ws = wb.active
    ws.title = "KRT_APP.1.1.xlsx"
    # Header row
    ws["A1"] = "APP.1.1"
    ws["B1"] = "Name"
    ws["C1"] = "CIA"
    ws["D1"] = "G 0.14"
    ws["E1"] = "G 0.39"
    # APP.1.1.A1 (deprecated, no links)
    ws.append(["APP.1.1.A1", "ENTFALLEN", "-", "-", "-"])
    # APP.1.1.A2 addresses G 0.39, no protection goals
    ws.append(["APP.1.1.A2", "Einschränken", "-", "-", "X"])

    # Second sheet for ORP.1
    ws2 = wb.create_sheet("KRT_ORP.1.xlsx")
    ws2["A1"] = "ORP.1"
    ws2["B1"] = "Name"
    ws2["C1"] = "CIA"
    ws2["D1"] = "G 0.14"
    ws2["E1"] = "G 0.39"
    # ORP.1.A1 addresses G 0.14 with CI goals
    ws2.append(["ORP.1.A1", "Sicherheitsleitlinie", "CI", "X", "-"])
    wb.save(path)


@pytest.fixture
def synthetic_krt(tmp_path: Path) -> Path:
    path = tmp_path / "tiny_krt.xlsx"
    _build_synthetic_krt(path)
    return path


def test_imports_links_and_goals(tiny_db, synthetic_krt):
    # tiny_db fixture already has requirement_threat rows; the importer
    # wipes them and re-inserts from KRT.
    stats = import_krt(tiny_db, synthetic_krt)
    assert stats.sheets_processed == 2
    assert stats.links_inserted == 2  # APP.1.1.A2 -> G 0.39, ORP.1.A1 -> G 0.14
    assert stats.goal_markers_inserted == 2  # C + I on ORP.1.A1 -> G 0.14

    rows = tiny_db.execute("""
        SELECT r.code AS req, t.code AS threat,
               GROUP_CONCAT(g.goal, ',') AS goals
        FROM requirement_threat rt
        JOIN requirement r ON r.id = rt.requirement_id
        JOIN threat t      ON t.id = rt.threat_id
        LEFT JOIN requirement_threat_goal g
            ON g.requirement_threat_id = rt.id
        GROUP BY rt.id ORDER BY r.code
    """).fetchall()
    assert [dict(r) for r in rows] == [
        {"req": "APP.1.1.A2", "threat": "G 0.39", "goals": None},
        {"req": "ORP.1.A1", "threat": "G 0.14", "goals": "C,I"},
    ]


def test_unknown_requirement_is_counted_not_raised(tiny_db, tmp_path: Path):
    """A KRT row pointing at a requirement that doesn't exist must not crash."""
    import openpyxl

    p = tmp_path / "krt_with_unknown.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KRT_APP.1.1.xlsx"
    ws["A1"] = "APP.1.1"
    ws["B1"] = "Name"
    ws["C1"] = "CIA"
    ws["D1"] = "G 0.39"
    ws.append(["APP.99.A1", "Ghost", "-", "X"])  # requirement unknown
    ws.append(["APP.1.1.A2", "Einschränken", "-", "X"])  # known
    wb.save(p)

    stats = import_krt(tiny_db, p)
    assert stats.requirements_unknown == 1
    assert stats.links_inserted == 1


def test_reimport_wipes_old_links(tiny_db, synthetic_krt):
    first = import_krt(tiny_db, synthetic_krt)
    second = import_krt(tiny_db, synthetic_krt)
    assert first.links_inserted == second.links_inserted
    n = tiny_db.execute("SELECT COUNT(*) FROM requirement_threat").fetchone()[0]
    assert n == second.links_inserted


def test_missing_catalog_raises(tmp_path: Path, synthetic_krt):
    """If no catalog row exists, the importer refuses to run."""
    from grundschutz_mcp.db import apply_schema, connect

    conn = connect(tmp_path / "empty.db")
    apply_schema(conn)
    with pytest.raises(RuntimeError, match="No catalog row"):
        import_krt(conn, synthetic_krt)


# ---------------------------------------------------------------------------
# Smoketest against the real KRT (skipped when not fetched)
# ---------------------------------------------------------------------------


@pytest.mark.skipif(
    not REAL_KRT.is_file(),
    reason="krt2023.xlsx missing - run `make fetch-krt` to enable.",
)
def test_smoketest_against_real_krt(tmp_path: Path):
    """Build the full catalog from XML, then import the real KRT into it."""
    from grundschutz_mcp.db import apply_schema, connect
    from grundschutz_mcp.ingest.xml_importer import import_xml

    xml = Path(__file__).resolve().parents[1] / "XML_Kompendium_2023.xml"
    if not xml.is_file():
        pytest.skip("XML_Kompendium_2023.xml missing")

    conn = connect(tmp_path / "real.db")
    apply_schema(conn)
    import_xml(conn, xml)
    stats = import_krt(conn, REAL_KRT)

    # Hard counts the 2023 catalog must produce.
    assert stats.sheets_processed == 111
    assert stats.links_inserted == 5888
    assert stats.goal_markers_inserted == 2820
    assert stats.requirements_unknown == 0
    assert stats.threats_unknown == 0
    conn.close()
