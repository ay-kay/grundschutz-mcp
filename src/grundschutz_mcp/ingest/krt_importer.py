"""Import the BSI Kreuzreferenztabelle (KRT) into the database.

The KRT is the M:N mapping between Anforderungen and elementare
Gefährdungen plus the per-requirement protection goals (C/I/A). It is
the one piece of data that is NOT in the DocBook XML; BSI publishes it
as a separate XLSX.

Source URL (fetched by ``scripts/fetch-bsi-krt.sh``):
    https://www.bsi.bund.de/.../krt2023_Excel.xlsx

File layout (empirically observed against the 2023 release):
    one sheet per Baustein, sheet name == ``KRT_<Code>.xlsx``
    row 1, col A : Baustein code
    row 1, col B : the literal string "Name"
    row 1, col C : the literal string "CIA"
    row 1, col D..: header cells with elementary threat codes
                    (``"G 0.14"``, ``"G 0.18"``, ...)
    rows 2..    : one Anforderung per row
        col A: requirement code (e.g. ``"APP.1.1.A2"``)
        col B: requirement title
        col C: CIA string (``"C"``, ``"I"``, ``"A"`` or combinations like
               ``"CI"``, ``"CIA"``; ``"-"`` / empty when the requirement
               has no protection-goal markers)
        col D..: ``"X"`` if the requirement addresses the threat in that
                 column, ``"-"`` / empty otherwise

The importer requires that the schema is in place AND that the XML
structure importer has already created the ``requirement`` and ``threat``
rows. KRT only adds the M:N links, never source entities.

Idempotency: existing ``requirement_threat`` rows for the catalog's
requirements are wiped at the start. The cascading FK delete also
removes ``requirement_threat_goal`` entries.
"""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from pathlib import Path

import structlog

log = structlog.get_logger(__name__)


@dataclass
class KrtImportStats:
    sheets_processed: int = 0
    links_inserted: int = 0
    goal_markers_inserted: int = 0
    requirements_unknown: int = 0  # in KRT but not in DB
    threats_unknown: int = 0  # in KRT but not in DB

    def as_dict(self) -> dict[str, int]:
        return dict(self.__dict__)


def import_krt(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    catalog_name: str = "BSI IT-Grundschutz 200-2",
    edition: str = "2023",
) -> KrtImportStats:
    """Load the BSI KRT XLSX into ``requirement_threat`` + ``_threat_goal``.

    Returns counts of every entity touched. Runs inside one transaction;
    if anything raises, the database is left unchanged.
    """
    # Heavy import only when actually needed.
    import openpyxl

    catalog_row = conn.execute(
        "SELECT id FROM catalog WHERE name = ? AND edition = ?",
        (catalog_name, edition),
    ).fetchone()
    if catalog_row is None:
        raise RuntimeError(
            f"No catalog row for name={catalog_name!r} edition={edition!r}; "
            f"run the XML structure importer first."
        )
    catalog_id = catalog_row["id"]

    # Build lookup tables (code -> id) so we don't fire one SELECT per cell.
    req_id_by_code = {
        r["code"]: r["id"]
        for r in conn.execute(
            "SELECT id, code FROM requirement WHERE catalog_id = ?",
            (catalog_id,),
        )
    }
    threat_id_by_code = {
        r["code"]: r["id"]
        for r in conn.execute(
            "SELECT id, code FROM threat WHERE catalog_id = ?",
            (catalog_id,),
        )
    }

    stats = KrtImportStats()
    try:
        with conn:
            _clear_existing_links(conn, catalog_id)
            wb = openpyxl.load_workbook(
                str(Path(xlsx_path)),
                data_only=True,
                read_only=False,
            )
            try:
                for sheet_name in wb.sheetnames:
                    _process_sheet(
                        conn=conn,
                        ws=wb[sheet_name],
                        req_id_by_code=req_id_by_code,
                        threat_id_by_code=threat_id_by_code,
                        stats=stats,
                    )
                    stats.sheets_processed += 1
            finally:
                wb.close()
    except Exception:
        log.exception("krt_import_failed")
        raise

    log.info("krt_import_done", **stats.as_dict())
    return stats


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------


def _clear_existing_links(conn: sqlite3.Connection, catalog_id: int) -> None:
    """Wipe requirement_threat rows for this catalog (cascades to goals)."""
    conn.execute(
        "DELETE FROM requirement_threat WHERE requirement_id IN ("
        "  SELECT id FROM requirement WHERE catalog_id = ?"
        ")",
        (catalog_id,),
    )


def _process_sheet(
    *,
    conn: sqlite3.Connection,
    ws,  # openpyxl Worksheet
    req_id_by_code: dict[str, int],
    threat_id_by_code: dict[str, int],
    stats: KrtImportStats,
) -> None:
    """Walk one sheet, INSERT the requirement_threat + goal rows."""
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return

    # Map: column index -> threat code. Columns from index 3 onward
    # carry threat codes; intermediate empty cells (rare) are skipped.
    threat_cols: list[tuple[int, str]] = []
    for col_idx, cell in enumerate(header[3:], start=3):
        if cell is None:
            continue
        code = str(cell).strip()
        if not code.startswith("G "):
            continue
        threat_cols.append((col_idx, code))

    for row in rows_iter:
        if not row or row[0] is None:
            continue
        req_code = str(row[0]).strip()
        if not req_code:
            continue
        req_id = req_id_by_code.get(req_code)
        if req_id is None:
            stats.requirements_unknown += 1
            continue

        # Protection-goal string for this requirement: applies to all
        # links emitted in this row.
        goals = _parse_cia(row[2] if len(row) > 2 else None)

        for col_idx, threat_code in threat_cols:
            if col_idx >= len(row):
                continue
            cell = row[col_idx]
            if not _is_mark(cell):
                continue
            threat_id = threat_id_by_code.get(threat_code)
            if threat_id is None:
                stats.threats_unknown += 1
                continue

            cur = conn.execute(
                "INSERT INTO requirement_threat (requirement_id, threat_id) "
                "VALUES (?, ?) "
                "ON CONFLICT (requirement_id, threat_id) DO NOTHING",
                (req_id, threat_id),
            )
            if cur.rowcount > 0:
                stats.links_inserted += 1

            # Get the link id (insert may have skipped due to ON CONFLICT).
            link = conn.execute(
                "SELECT id FROM requirement_threat WHERE requirement_id = ? AND threat_id = ?",
                (req_id, threat_id),
            ).fetchone()
            link_id = link["id"]

            for goal in goals:
                conn.execute(
                    "INSERT OR IGNORE INTO requirement_threat_goal "
                    "(requirement_threat_id, goal) VALUES (?, ?)",
                    (link_id, goal),
                )
                # We don't bump goal_markers_inserted on conflict; rough
                # count via DELETE-then-INSERT pattern means this is the
                # actual fresh count.
                stats.goal_markers_inserted += 1


def _is_mark(cell) -> bool:
    """Whether a KRT body cell denotes "requirement addresses this threat"."""
    if cell is None:
        return False
    return str(cell).strip().upper() == "X"


def _parse_cia(cell) -> list[str]:
    """Parse the CIA-column string into a list of distinct goals."""
    if cell is None:
        return []
    text = str(cell).strip().upper()
    if not text or text == "-":
        return []
    return [ch for ch in ("C", "I", "A") if ch in text]
